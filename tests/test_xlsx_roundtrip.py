import openpyxl

from doctranslate.pipeline import run_pipeline


def test_xlsx_roundtrip(tmp_path, fake_llm_client):
    src = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "Role"
    ws["A2"] = "Alice"
    ws["B2"] = "Engineer"
    ws["C2"] = 42
    wb.save(src)

    out_path = run_pipeline(str(src), "en", "th", fake_llm_client)

    assert out_path.exists()
    out_wb = openpyxl.load_workbook(out_path)
    out_ws = out_wb.active

    assert out_ws["A1"].value == "[TR]Name"
    assert out_ws["B1"].value == "[TR]Role"
    assert out_ws["A2"].value == "[TR]Alice"
    assert out_ws["B2"].value == "[TR]Engineer"
    assert out_ws["C2"].value == 42


def test_xlsx_roundtrip_skips_formulas(tmp_path, fake_llm_client):
    src = tmp_path / "with_formula.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Total"
    ws["A2"] = "=SUM(1,2)"
    wb.save(src)

    out_path = run_pipeline(str(src), "en", "th", fake_llm_client)

    out_wb = openpyxl.load_workbook(out_path)
    out_ws = out_wb.active
    assert out_ws["A1"].value == "[TR]Total"
    assert out_ws["A2"].value == "=SUM(1,2)"
